from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# choose file 
# report = input(enter path to this weeks report:)
# report = 'sample_report.xlsx'

wb = load_workbook('sample_report.xlsx')
ws = wb.active


for row in range(1, 102):
    print(ws['D' + str(row)].value)